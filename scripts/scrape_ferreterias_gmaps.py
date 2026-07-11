# -*- coding: utf-8 -*-
"""
Baja TODAS las ferreterías de Google Maps alrededor del depósito de Uniproveedores
(Bacacay 4726, Comuna 10, CABA) y las exporta a un Excel ordenado por distancia.

Esta es la forma REAL de tener el listado completo: Google Maps es la base más
actualizada y grande. El script recorre la zona en una grilla de puntos y, en cada
punto, pide las ferreterías cercanas, junta todo, elimina duplicados y ordena por
cercanía al depósito.

----------------------------------------------------------------------
CÓMO USARLO (una sola vez, ~10 min de setup)
----------------------------------------------------------------------
1) Conseguir una API key de Google Maps (gratis hasta un límite mensual alto):
   - Entrá a https://console.cloud.google.com/  → creá un proyecto.
   - Menú "APIs y servicios" → "Biblioteca" → activá **Places API** y **Geocoding API**.
   - "Credenciales" → "Crear credenciales" → "Clave de API". Copiá la clave.
   - (Google da USD 200/mes de crédito gratis; este barrido entra holgado en ese margen.)

2) Instalar dependencias (una vez):
       pip install requests openpyxl

3) Correr:
       export GOOGLE_MAPS_API_KEY="TU_CLAVE"
       python3 scripts/scrape_ferreterias_gmaps.py

   Opciones:
       RADIO_KM=6   → radio de barrido en km (default 6, como pediste)
       Salida: data/Ferreterias-GoogleMaps-CABA.xlsx

----------------------------------------------------------------------
NOTA: sin API key el script NO corre (no inventa datos). Te lo dice y sale.
----------------------------------------------------------------------
"""
import os
import time
import math
import sys

try:
    import requests
except ImportError:
    sys.exit("Falta 'requests'. Instalá con:  pip install requests openpyxl")

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

API_KEY = os.environ.get("GOOGLE_MAPS_API_KEY", "").strip()
DEPOSITO = "Bacacay 4726, Ciudad Autónoma de Buenos Aires, Argentina"
RADIO_KM = float(os.environ.get("RADIO_KM", "6"))
KEYWORDS = ["ferretería", "ferretería industrial", "herramientas", "corralón"]
OUT = "data/Ferreterias-GoogleMaps-CABA.xlsx"

VERDE = "A4D72B"; NEGRO = "0D0D0D"


def die(msg):
    print("\n" + msg + "\n")
    sys.exit(1)


if not API_KEY:
    die("No hay GOOGLE_MAPS_API_KEY. Seguí los pasos del encabezado de este archivo "
        "para obtener la clave y luego:\n"
        '   export GOOGLE_MAPS_API_KEY="TU_CLAVE"\n'
        "   python3 scripts/scrape_ferreterias_gmaps.py")


def geocode(address):
    r = requests.get("https://maps.googleapis.com/maps/api/geocode/json",
                     params={"address": address, "key": API_KEY}, timeout=30).json()
    if r.get("status") != "OK":
        die(f"Geocoding falló ({r.get('status')}). Revisá que Geocoding API esté activada y la key sea válida.")
    loc = r["results"][0]["geometry"]["location"]
    return loc["lat"], loc["lng"]


def haversine_km(lat1, lon1, lat2, lon2):
    R = 6371.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dl = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * R * math.asin(math.sqrt(a))


def grid_points(clat, clon, radio_km, step_km=0.7):
    """Puntos de una grilla que cubre el círculo de radio radio_km."""
    pts = []
    dlat = step_km / 110.574
    dlon = step_km / (111.320 * math.cos(math.radians(clat)))
    n = int(radio_km / step_km) + 1
    for i in range(-n, n + 1):
        for j in range(-n, n + 1):
            lat = clat + i * dlat
            lon = clon + j * dlon
            if haversine_km(clat, clon, lat, lon) <= radio_km:
                pts.append((lat, lon))
    return pts


def nearby(lat, lon, keyword, radius_m=700):
    """Places Nearby Search con paginación (hasta 3 páginas = 60 resultados)."""
    out = []
    params = {"location": f"{lat},{lon}", "radius": radius_m,
              "keyword": keyword, "language": "es", "key": API_KEY}
    url = "https://maps.googleapis.com/maps/api/place/nearbysearch/json"
    for _ in range(3):
        r = requests.get(url, params=params, timeout=30).json()
        st = r.get("status")
        if st not in ("OK", "ZERO_RESULTS"):
            print(f"  aviso: status {st}")
            break
        out.extend(r.get("results", []))
        token = r.get("next_page_token")
        if not token:
            break
        time.sleep(2)  # el token tarda unos segundos en activarse
        params = {"pagetoken": token, "key": API_KEY}
    return out


def details(place_id):
    fields = "name,formatted_address,formatted_phone_number,international_phone_number,website,opening_hours,rating,user_ratings_total,url"
    r = requests.get("https://maps.googleapis.com/maps/api/place/details/json",
                     params={"place_id": place_id, "fields": fields,
                             "language": "es", "key": API_KEY}, timeout=30).json()
    return r.get("result", {}) if r.get("status") == "OK" else {}


def main():
    print(f"Depósito: {DEPOSITO}")
    clat, clon = geocode(DEPOSITO)
    print(f"Coordenadas: {clat:.5f}, {clon:.5f}  |  radio {RADIO_KM} km")

    pts = grid_points(clat, clon, RADIO_KM)
    print(f"Barriendo {len(pts)} puntos x {len(KEYWORDS)} términos...")

    seen = {}  # place_id -> result básico
    for k, (lat, lon) in enumerate(pts, 1):
        for kw in KEYWORDS:
            for res in nearby(lat, lon, kw):
                pid = res.get("place_id")
                if pid and pid not in seen:
                    seen[pid] = res
        if k % 10 == 0:
            print(f"  {k}/{len(pts)} puntos · {len(seen)} lugares únicos")

    print(f"Total lugares únicos: {len(seen)}. Pidiendo detalles...")

    rows = []
    for i, (pid, res) in enumerate(seen.items(), 1):
        d = details(pid)
        loc = res.get("geometry", {}).get("location", {})
        dist = haversine_km(clat, clon, loc.get("lat", clat), loc.get("lng", clon))
        hours = "; ".join(d.get("opening_hours", {}).get("weekday_text", [])) if d.get("opening_hours") else ""
        rows.append({
            "nombre": d.get("name") or res.get("name", ""),
            "direccion": d.get("formatted_address", res.get("vicinity", "")),
            "tel": d.get("formatted_phone_number", ""),
            "internacional": d.get("international_phone_number", ""),
            "web": d.get("website", ""),
            "horarios": hours,
            "rating": d.get("rating", ""),
            "reviews": d.get("user_ratings_total", ""),
            "maps": d.get("url", ""),
            "dist_km": round(dist, 2),
        })
        if i % 25 == 0:
            print(f"  detalles {i}/{len(seen)}")

    rows.sort(key=lambda x: x["dist_km"])
    export(rows, clat, clon)
    print(f"\nOK -> {OUT}  ({len(rows)} ferreterías, ordenadas por distancia al depósito)")


def export(rows, clat, clon):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ferreterías GMaps"
    cols = ["#", "Dist. km", "Nombre", "Dirección", "Teléfono", "WhatsApp (intl.)",
            "Sitio web", "Horarios", "Rating", "Reseñas", "Ver en Maps",
            "Email", "Encargado", "Estado", "Prioridad", "Notas"]
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.merge_cells("A1:P1")
    c = ws["A1"]
    c.value = f"UNIPROVEEDORES · Ferreterías desde Google Maps (radio {RADIO_KM} km del depósito) — {len(rows)} resultados"
    c.font = Font(bold=True, color="FFFFFF", size=14)
    c.fill = PatternFill("solid", fgColor=NEGRO); ws.row_dimensions[1].height = 24
    for j, name in enumerate(cols, 1):
        cell = ws.cell(row=2, column=j, value=name)
        cell.font = Font(bold=True, color="0D0D0D"); cell.fill = PatternFill("solid", fgColor=VERDE)
        cell.alignment = Alignment(horizontal="center", vertical="center"); cell.border = border
    for i, x in enumerate(rows, 1):
        r = i + 2
        vals = [i, x["dist_km"], x["nombre"], x["direccion"], x["tel"], x["internacional"],
                x["web"], x["horarios"], x["rating"], x["reviews"], x["maps"],
                "", "", "Sin contactar", "", ""]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=j, value=v)
            cell.alignment = Alignment(wrap_text=True, vertical="top"); cell.border = border
    for j, w in enumerate([4, 8, 30, 34, 15, 17, 24, 30, 8, 9, 16, 20, 16, 14, 9, 24], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:P{len(rows)+2}"
    os.makedirs("data", exist_ok=True)
    wb.save(OUT)


if __name__ == "__main__":
    main()
