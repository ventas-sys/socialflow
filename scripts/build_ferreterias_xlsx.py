# -*- coding: utf-8 -*-
"""Genera el Excel de prospeccion de ferreterias CABA + plan de redes para Uniproveedores.

- Ordenado por CERCANIA al deposito (Bacacay 4726, Comuna 10 - borde Floresta / Velez Sarsfield).
  Anillo 1 = Comuna 10 e inmediato (~0-2 km). Anillo 2 = ~2-6 km. Anillo 3 = resto de CABA.
- Datos de contacto: directorios publicos (guiaferreterias, paginasamarillas, argentino.com.ar,
  buscabaires, licuo, IG/FB, webs propias). Campos no publicos (encargado, email, whatsapp
  directo, horario exacto) quedan vacios y se completan en la prospeccion. NO se inventan.
- Para el listado COMPLETO de Google Maps: usar scripts/scrape_ferreterias_gmaps.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

VERDE = "A4D72B"; NEGRO = "0D0D0D"; GRIS = "9AA0A6"

wb = openpyxl.Workbook()
thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_font = Font(bold=True, color="0D0D0D", size=11)
hdr_fill = PatternFill("solid", fgColor=VERDE)
title_font = Font(bold=True, color="FFFFFF", size=15)
title_fill = PatternFill("solid", fgColor=NEGRO)
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center")

# Fills por anillo
FILL_A1 = PatternFill("solid", fgColor="E9F6C8")  # verde muy claro = mas cerca
FILL_A2 = PatternFill("solid", fgColor="F6F6E3")
FILL_A3 = PatternFill("solid", fgColor="FFFFFF")

# =====================================================================
# HOJA 1 - FERRETERIAS CABA
# =====================================================================
ws = wb.active
ws.title = "Ferreterías CABA"
cols = ["#", "Anillo", "Nombre", "Barrio", "Dirección", "Teléfono", "WhatsApp",
        "Email", "Sitio web / Redes", "Instagram", "Horarios", "Encargado",
        "Estado", "Prioridad", "Notas / Fuente"]

ws.merge_cells("A1:O1")
c = ws["A1"]
c.value = "UNIPROVEEDORES · Ferreterías CABA — ordenadas por cercanía al depósito (Bacacay 4726, Comuna 10)"
c.font = title_font; c.fill = title_fill
c.alignment = Alignment(horizontal="left", vertical="center"); ws.row_dimensions[1].height = 26

ws.merge_cells("A2:O2")
c = ws["A2"]
c.value = ("Anillo 1 = Comuna 10 e inmediato (~0-2 km). Anillo 2 = ~2-6 km. Anillo 3 = resto de CABA. "
           "Empezá por Anillo 1 (verde). Campos vacíos = completar en el contacto (ver hoja 'Cómo completar y escalar'). "
           "Para bajar TODAS las de Google Maps automáticamente: scripts/scrape_ferreterias_gmaps.py")
c.font = Font(italic=True, color="666666", size=9); c.alignment = wrap
ws.row_dimensions[2].height = 30

for j, name in enumerate(cols, start=1):
    cell = ws.cell(row=3, column=j, value=name)
    cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center; cell.border = border

# (anillo, nombre, barrio, direccion, tel, whatsapp, email, web/redes, instagram, horarios, encargado, estado, prioridad, notas)
D = [
 # ---------- ANILLO 1 (Comuna 10 e inmediato, ~0-2 km) ----------
 (1,"Ferretería (Belaústegui)","Floresta","Dr. Luis Belaústegui 3901","011 4639-0060","confirmar","","","","","","Sin contactar","Alta","Muy cerca del depósito."),
 (1,"Ferretería (Yerbal)","Floresta","Yerbal 3920","011 4672-2218","confirmar","","","","","","Sin contactar","Alta","Muy cerca del depósito."),
 (1,"Ferretería (Gaona 4277)","Floresta","Av. Gaona 4277","011 4671-7132","confirmar","","","","","","Sin contactar","Alta","Sobre avenida, buen tránsito."),
 (1,"Ferretería (J. V. González)","Floresta","Joaquín V. González 19","011 4674-6435","confirmar","","","","","","Sin contactar","Alta","Muy cerca del depósito."),
 (1,"Ferretería (Lacarra 53)","Vélez Sarsfield","Av. Lacarra 53","011 4674-1101","confirmar","","","","","","Sin contactar","Alta","Barrio del depósito."),
 (1,"Ferretería (E. de San Martín)","Vélez Sarsfield","E. de San Martín 4944","011 4674-0327","confirmar","","","","","","Sin contactar","Alta","Barrio del depósito."),
 (1,"Ferretería (Rivadavia 9825)","Villa Luro","Av. Rivadavia 9825","011 4682-1993","confirmar","","","","","","Sin contactar","Alta","Sobre Rivadavia."),
 (1,"Ferretería (Víctor Hugo)","Villa Luro","Víctor Hugo 43","011 4682-8050","confirmar","","","","","","Sin contactar","Alta","Cerca del depósito."),
 (1,"Ferretería (Larrañaga)","Villa Luro","D. Larrañaga 694","011 4683-1181","confirmar","","","","","","Sin contactar","Media","Cerca del depósito."),
 (1,"Ferretería (Álvarez Jonte 4663)","Monte Castro","Av. Álvarez Jonte 4663","011 4567-7997","confirmar","","","","","","Sin contactar","Alta","Sobre avenida."),
 (1,"Ferretería (Benito Juárez)","Monte Castro","Benito Juárez 1855","","confirmar","","","","","","Sin contactar","Media","Directorio Monte Castro."),
 (1,"Ferretería (Lope de Vega 2240)","Villa Real","Av. Lope de Vega 2240","011 4568-8186","confirmar","","","","","","Sin contactar","Media","Sobre avenida."),
 (1,"Ferretería (Lope de Vega 2832)","Villa Real","Av. Lope de Vega 2832","011 4568-8402","confirmar","","","","","","Sin contactar","Media","Sobre avenida."),
 (1,"Ferretería (Nazca 1311)","Villa Santa Rita","Av. Nazca 1311","011 4584-2809","confirmar","","","","","","Sin contactar","Alta","Sobre Nazca (avenida clave)."),
 (1,"Ferretería (Nazca 1826)","Villa Santa Rita","Av. Nazca 1826","011 4584-6513","confirmar","","","","","","Sin contactar","Alta","Sobre Nazca."),
 (1,"Materiales Cardinale S.R.L.","Villa del Parque","Av. Mosconi 3992","","confirmar","","","","","","Sin contactar","Alta","Materiales + ferretería. Volumen."),
 (1,"Ferretería del Parque","Villa del Parque","(consultar - Páginas Amarillas)","","confirmar","","","","","","Sin contactar","Media","Ficha PáginasAmarillas."),
 (1,"Ferretería (Nazarre)","Villa del Parque","Nazarre 3256","011 4587-6243","confirmar","","","","","","Sin contactar","Media","Directorio V. del Parque."),
 (1,"Herramientas Dali S.R.L.","Villa del Parque","Av. Juan B. Justo 5768","","confirmar","","","","","","Sin contactar","Alta","Herramientas. Posible mayorista."),

 # ---------- ANILLO 2 (~2-6 km) ----------
 (2,"Ferretería Santa María","Caballito","Av. La Plata 213","","confirmar","","ferreteriasantamaria.com.ar","","","","Sin contactar","Alta","Web propia. Buen prospecto."),
 (2,"Ferretería Online","Caballito","Av. Rivadavia 5456 Local 16","","confirmar","","ferreteriaonline.com.ar","","","","Sin contactar","Alta","E-commerce de herramientas industriales."),
 (2,"LB Ferreterías","Caballito","Av. Honorio Pueyrredón 1182","","confirmar","","lbferreterias.com.ar","","","","Sin contactar","Alta","Web propia."),
 (2,"Mercado del Progreso (sector ferret.)","Caballito","Av. Rivadavia 5430","","confirmar","","mercadodelprogreso.com","","","","Sin contactar","Media","Puestos de ferretería/electricidad."),
 (2,"Ferretería (Avellaneda 2287)","Flores","Av. Avellaneda 2287","011 4631-0351","confirmar","","","","","","Sin contactar","Media","Sobre avenida comercial."),
 (2,"Ferretería (Directorio 2320)","Flores","Av. Directorio 2320","011 4634-1301","confirmar","","","","","","Sin contactar","Media","Directorio Flores."),
 (2,"Ferretería (Eva Perón 2951)","Flores","Av. Eva Perón 2951","011 4660-6105","confirmar","","","","","","Sin contactar","Media","Sobre avenida."),
 (2,"Argentec Herramientas (Flores)","Flores","Av. Gaona 3059","","confirmar","","argentec.com.ar","","","","Sin contactar","Alta","Cadena con red de sucursales y web."),
 (2,"Ferretería (Eva Perón 4285)","Parque Avellaneda","Av. Eva Perón 4285","011 4671-6241","confirmar","","","","","","Sin contactar","Media","Sobre avenida."),
 (2,"Ferretería (Mariano Acosta)","Parque Avellaneda","Mariano Acosta 1066","011 4613-9083","confirmar","","","","","","Sin contactar","Media","Directorio Pque Avellaneda."),
 (2,"Ferretería (Eva Perón 4388)","Parque Avellaneda","Av. Eva Perón 4388","011 4672-1745","confirmar","","","","","","Sin contactar","Media","Sobre avenida."),
 (2,"Ferretería (Barco Centenera)","Parque Chacabuco","Av. del Barco Centenera 2227","011 4923-4098","confirmar","","","","","","Sin contactar","Media","Sobre avenida."),
 (2,"Ferretería (Directorio 1650)","Parque Chacabuco","Av. Directorio 1650","011 4432-79 (confirmar)","confirmar","","","","","","Sin contactar","Media","Verificar teléfono completo."),
 (2,"Ferretería (Donato Álvarez 2203)","La Paternal","Av. Donato Álvarez 2203","011 4826-4009","confirmar","","","","","","Sin contactar","Media","Sobre avenida."),
 (2,"Ferretería (Donato Álvarez 3156)","La Paternal","Av. Donato Álvarez 3156","011 4521-2682","confirmar","","","","","","Sin contactar","Media","Sobre avenida."),
 (2,"Ferretería (San Martín 3202)","La Paternal","Av. San Martín 3202","011 4581-6195","confirmar","","","","","","Sin contactar","Media","Sobre avenida San Martín."),
 (2,"Argentec Herramientas (V. Crespo)","Villa Crespo","Av. Córdoba 5786","","confirmar","","argentec.com.ar","","","","Sin contactar","Alta","Cadena con web."),
 (2,"Argentec Herramientas (V. Crespo 2)","Villa Crespo","Av. Córdoba 4038","","confirmar","","argentec.com.ar","","","","Sin contactar","Media","Cadena con web."),
 (2,"Herramientas Dali S.R.L. (V. Crespo)","Villa Crespo","Montenegro 32","","confirmar","","","","","","Sin contactar","Media","Directorio V. Crespo."),
 (2,"Argentec Herramientas (Boedo)","Boedo","Av. Asamblea 524","","confirmar","","argentec.com.ar","","","","Sin contactar","Alta","Cadena con web."),
 (2,"Casa Aldo (ferretería y herrajes)","Villa Pueyrredón","Salvador M. del Carril 2921","011 4572-7412","confirmar","","","","","","Sin contactar","Media","Ferretería + herrajes."),
 (2,"Ferretería (Artigas 5676)","Villa Pueyrredón","Gral. Artigas 5676","011 4572-7616","confirmar","","","","","","Sin contactar","Baja","Directorio V. Pueyrredón."),
 (2,"Ferretería (Artigas 5299)","Villa Pueyrredón","Gral. Artigas 5299","011 4571-2586","confirmar","","","","","","Sin contactar","Baja","Directorio V. Pueyrredón."),
 (2,"Ferretería (Alberdi 5775)","Mataderos","Av. J.B. Alberdi 5775","","confirmar","","","","","","Sin contactar","Media","Sobre avenida."),
 (2,"Ferretería (Larrazábal)","Mataderos","Av. Larrazábal 966","","confirmar","","","","","","Sin contactar","Media","Sobre avenida."),
 (2,"Ferretería (Eva Perón 5553)","Mataderos","Av. Eva Perón 5553","","confirmar","","","","","","Sin contactar","Media","Sobre avenida."),
 (2,"Ferretería (Oliden)","Mataderos","Oliden 2150","","confirmar","","","","","","Sin contactar","Baja","Directorio Mataderos."),
 (2,"Ferretería Avenida","Mataderos","(consultar - Buscabaires)","","confirmar","","","","","","Sin contactar","Baja","Verificar dirección."),
 (2,"Ferretería Montiel","Liniers","(consultar - Infoisinfo)","","confirmar","","","","","","Sin contactar","Baja","Verificar dirección."),
 (2,"Ferretería El Ahorro","Liniers","(consultar - Infoisinfo)","","confirmar","","","","","","Sin contactar","Baja","Verificar dirección."),
 (2,"Ferretería (Sanabria)","Villa Devoto","Sanabria 2885","","confirmar","","","","","","Sin contactar","Media","Directorio V. Devoto."),
 (2,"Ferretería (Segurola)","Villa Devoto","Av. Segurola 2884","","confirmar","","","","","","Sin contactar","Media","Sobre avenida."),
 (2,"Ferretería (Marcos Paz)","Villa Devoto","Marcos Paz 4510","","confirmar","","","","","","Sin contactar","Media","Directorio V. Devoto."),

 # ---------- ANILLO 3 (resto de CABA, >6 km) ----------
 (3,"Ferretería del Once","Balvanera / Once","Sarmiento 2846","011 4687-0696","confirmar","","facebook.com/p/Ferreteria-del-once","","L-V 9-18, Sáb AM","","Sin contactar","Alta","Zona mayorista Once. 35+ años."),
 (3,"RIVAFER","Once / Balvanera","(consultar web)","","","","rivafer.com.ar","","","","Sin contactar","Alta","E-commerce propio."),
 (3,"Ferretería El Once","Once / Balvanera","(tienda online)","","","","ferreteriaelonce.com","","","","Sin contactar","Alta","Vende online. Mayorista."),
 (3,"MAABA Distribuidora","Villa Lugano","Zelada 6523","011 7569-4425","11 5429-1284","","maaba.com.ar","","","","Sin contactar","Alta","Distribuidora. Móvil = WhatsApp probable."),
 (3,"Ferretería Belgrano","Belgrano","Manuel Ugarte 2391 / Zapiola 2468","","confirmar","","","instagram.com/ferreteriabelgrano2","","","Sin contactar","Alta","2 sucursales. IG activo."),
 (3,"Ferretería (Monroe 3626)","Belgrano","Av. Monroe 3626","011 4541-6047","confirmar","","","","","","Sin contactar","Media","Directorio Belgrano."),
 (3,"Ferretería (Juramento)","Belgrano","Juramento 3125","011 4544-9991","confirmar","","","","","","Sin contactar","Media","Directorio Belgrano."),
 (3,"Ferretería (Amenábar)","Belgrano","Amenábar 2273","011 4787-6355","confirmar","","","","","","Sin contactar","Media","Directorio Belgrano."),
 (3,"Ferretería (Blanco Encalada)","Belgrano","Blanco Encalada 2606","011 4787-4432","confirmar","","","","","","Sin contactar","Media","Directorio Belgrano."),
 (3,"Ferretería (Moldes)","Belgrano","Moldes 1802","011 4787-4316","confirmar","","","","","","Sin contactar","Media","Directorio Belgrano."),
 (3,"Ferretería y Cerraj. 24hs Trenque","Villa Urquiza","Echeverría 5099 (y Ávalos)","011 4521-9377","11 5002-4497","","","","24 hs","","Sin contactar","Alta","24hs. Móvil = WhatsApp probable."),
 (3,"Ferretería (Lavoisier)","Villa Urquiza","Lavoisier 3273","011 4573-2278","confirmar","","","","","","Sin contactar","Media","Directorio V. Urquiza."),
 (3,"Ferretería (Burela)","Villa Urquiza","Burela 2725","011 4521-4337","confirmar","","","","","","Sin contactar","Media","Directorio V. Urquiza."),
 (3,"Ferretería (Bauness)","Villa Urquiza","Bauness 2201","011 4521-9229","confirmar","","","","","","Sin contactar","Media","Directorio V. Urquiza."),
 (3,"Ferretería (Monroe 3897)","Villa Urquiza","Av. Monroe 3897","011 4545-4829","confirmar","","","","","","Sin contactar","Media","Directorio V. Urquiza."),
 (3,"Ferretería (Palermo)","Palermo","Jorge Luis Borges 2475 Local 37","","confirmar","","","","","","Sin contactar","Media","Directorio Palermo."),
 (3,"Ferretería (Crisólogo Larralde)","Núñez","Crisólogo Larralde 1982","","confirmar","","","","","","Sin contactar","Media","Directorio Núñez."),
 (3,"Ferretería (Pacheco)","Saavedra","Pacheco 3545","","confirmar","","","","","","Sin contactar","Media","Directorio Saavedra."),
 (3,"Ferretería (Ruiz Huidobro)","Saavedra","Ruiz Huidobro 2663","","confirmar","","","","","","Sin contactar","Media","Directorio Saavedra."),
 (3,"Ferretería San Telmo S.R.L.","San Telmo","(consultar IG/FB)","","confirmar","","facebook.com/p/FERRETERIA-SAN-TELMO-SRL","instagram.com/ferreteriasantelmo","","","Sin contactar","Alta","IG + FB activos."),
 (3,"Ferretería (Defensa)","San Telmo","Defensa 732","011 4362-3299","confirmar","","","","","","Sin contactar","Media","Directorio San Telmo."),
 (3,"Ferretería (Av. Brasil)","Constitución","Av. Brasil 601","011 4300-7763","confirmar","","","","","","Sin contactar","Media","Zona sur."),
 (3,"Ferretería (Chacabuco)","San Telmo","Chacabuco 798","011 4300-6970","confirmar","","","","","","Sin contactar","Media","Directorio San Telmo."),
 (3,"Ferretería (San Antonio)","Barracas","San Antonio 440","011 4302-6979","confirmar","","","","","","Sin contactar","Media","Directorio Barracas."),
 (3,"Ferretería (Salom)","Barracas","Salom 326","011 4302-5832","confirmar","","","","","","Sin contactar","Media","Directorio Barracas."),
 (3,"Ferretería (Rocha)","Barracas","Rocha 1600","011 4302-5122","confirmar","","","","","","Sin contactar","Media","Directorio Barracas."),
]

fills = {1: FILL_A1, 2: FILL_A2, 3: FILL_A3}
for i, row in enumerate(D, start=1):
    r = i + 3
    anillo = row[0]
    ws.cell(row=r, column=1, value=i).alignment = center
    ws.cell(row=r, column=1).border = border
    for j, val in enumerate(row, start=2):
        cell = ws.cell(row=r, column=j, value=(f"Anillo {val}" if j == 2 else val))
        cell.alignment = wrap; cell.border = border
        cell.fill = fills[anillo]
    ws.cell(row=r, column=1).fill = fills[anillo]

widths = [4, 8, 32, 17, 30, 16, 14, 8, 26, 24, 13, 14, 14, 9, 30]
for j, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(j)].width = w
ws.freeze_panes = "A4"
ws.auto_filter.ref = f"A3:O{len(D)+3}"

# =====================================================================
# HOJA 2 - COMO COMPLETAR Y ESCALAR
# =====================================================================
ws2 = wb.create_sheet("Cómo completar y escalar")
ws2.column_dimensions["A"].width = 3
ws2.column_dimensions["B"].width = 112

def block(ws, rows):
    r = 1
    for kind, text in rows:
        cell = ws.cell(row=r, column=2, value=text)
        if kind == "h1":
            cell.font = title_font; cell.fill = title_fill; ws.row_dimensions[r].height = 24
        elif kind == "h2":
            cell.font = Font(bold=True, color="0D0D0D", size=12); cell.fill = PatternFill("solid", fgColor=VERDE)
        elif kind == "b":
            cell.font = Font(bold=True, size=10); cell.alignment = wrap
        else:
            cell.font = Font(size=10); cell.alignment = wrap
        r += 1

block(ws2, [
 ("h1", "  Cómo completar datos que faltan y llegar a TODA CABA"),
 ("n", ""),
 ("h2", "  Realidad del número"),
 ("n", "No hay padrón público de ferreterías de CABA. Estimación realista: entre ~800 y ~1.500 ferreterías en toda la Ciudad "
        "(no 10.000: ese número sería para toda la Argentina). Igual son MUCHÍSIMAS más de las que podés cargar a mano → por eso está el script."),
 ("n", ""),
 ("h2", "  A) Bajar TODAS automáticamente (Google Maps) — recomendado"),
 ("b", "   scripts/scrape_ferreterias_gmaps.py  (ver instrucciones dentro del archivo)"),
 ("n", "   • Centra la búsqueda en el depósito (Bacacay 4726) y trae cada ferretería con nombre, dirección, teléfono, web,"),
 ("n", "     horario y rating, ordenadas por distancia. Exporta un Excel igual a este."),
 ("n", "   • Necesita una API key de Google Maps (Places API) — gratis hasta un límite mensual alto. Pasos en el script."),
 ("n", ""),
 ("h2", "  B) Completar Encargado / Email / WhatsApp / Horario de cada ficha"),
 ("n", "   1) Google Maps: la ficha da teléfono, web, horarios reales y link a IG/WhatsApp."),
 ("n", "   2) WhatsApp Business: abrí wa.me/54911XXXXXXXX; si tiene perfil de empresa, es WhatsApp válido → pintá la celda de verde."),
 ("n", "   3) El nombre del encargado se PREGUNTA en el primer contacto ('¿quién hace las compras?'). No se inventa."),
 ("n", ""),
 ("h2", "  C) Encontrar ferreterías NUEVAS cada semana"),
 ("n", "   • Google Maps: 'ferretería + barrio', mové el mapa y sumá las que falten."),
 ("n", "   • Directorios: guiaferreterias.com.ar, paginasamarillas.com.ar, argentino.com.ar, buscabaires.com, licuo.com.ar (por barrio)."),
 ("n", "   • Instagram/Facebook: 'ferretería + barrio' (muchas PyMEs solo están en redes)."),
 ("n", "   • Mercado Libre / El Ferretero (elferretero.com.ar): vendedores del rubro = posibles clientes mayoristas."),
 ("n", ""),
 ("h2", "  D) Estado de prospección"),
 ("n", "   Sin contactar → Contactado → Interesado → Cotización enviada → Cliente / Descartado. Filtrá 'Prioridad = Alta' primero."),
 ("n", ""),
 ("b", "  Regla de oro: no inventar datos. Un mail/WhatsApp equivocado quema el prospecto y la reputación de la marca."),
])

# =====================================================================
# HOJA 3 - PLAN SEMANAL REDES
# =====================================================================
ws3 = wb.create_sheet("Plan Semanal Redes")
for j, w in enumerate([14, 44, 44, 44], start=1):
    ws3.column_dimensions[get_column_letter(j)].width = w
ws3.merge_cells("A1:D1")
c = ws3["A1"]; c.value = "UNIPROVEEDORES · Plan Semanal — Redes + Prospección (empezar por el anillo cercano al depósito)"
c.font = title_font; c.fill = title_fill; ws3.row_dimensions[1].height = 24
for j, h in enumerate(["Día", "Prospección ferreterías", "Contenido / Redes", "Acción de crecimiento (seguidores)"], start=1):
    cell = ws3.cell(row=2, column=j, value=h)
    cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center; cell.border = border
plan = [
 ("Lunes","Elegí 1 barrio del Anillo 1/2 (empezá pegado al depósito). Cargá 10 ferreterías nuevas (o corré el script de Google Maps).",
  "IG+FB: producto estrella con precio (cápsula verde). X: oferta corta. Canal WhatsApp: 'ofertas de la semana'.",
  "Seguí 20 ferreterías del rubro en IG. Comentá en 5 posteos. Invitá contactos a seguir la Página FB."),
 ("Martes","Escribí por WhatsApp a 10 'Sin contactar' del Anillo 1. Completá Encargado y validá WhatsApp.",
  "YouTube: 1 Short (herramienta en acción, 20-40s). Mismo clip → Reel IG.",
  "YouTube: respondé todos los comentarios y pedí suscripción. IG: 3 stories con encuesta."),
 ("Miércoles","Cargá 10 ferreterías nuevas (Anillo 2). Pasá a 'Interesado' las que respondieron.",
  "LinkedIn: post B2B ('mayorista para ferreterías, envío y garantía'). IG: carrusel de catálogo.",
  "LinkedIn: conectá con 20 dueños/encargados de ferreterías. IG: collab con 1 proveedor."),
 ("Jueves","Enviá lista de precios/cotización a los 'Interesados'. Registrá en Notas.",
  "IG+FB: testimonio o 'antes/después'. Canal WhatsApp: novedad de stock. X: hilo con 3 tips.",
  "Concurso: 'seguinos + etiquetá a 2 ferreteros' por un descuento. Reglas en story fija."),
 ("Viernes","Seguimiento a todos los contactados de la semana. Cerrá lo posible a 'Cliente'.",
  "YouTube: video/Short de oferta de fin de semana. Reutilizá para IG/FB/TikTok/X.",
  "Compartí el video en TODAS las redes + Canal WhatsApp. Fijá el mejor post."),
 ("Sáb/Dom","Actualizá KPIs (leads nuevos, contactados, clientes). Planificá el barrio del lunes.",
  "Programá 2-3 posteos livianos. Respondé mensajes pendientes.",
  "Mirá qué post creció más → repetí ese formato la semana próxima."),
]
r = 3
for day, a, b, cc in plan:
    ws3.cell(row=r, column=1, value=day).font = Font(bold=True)
    for j, v in enumerate([a, b, cc], start=2):
        ws3.cell(row=r, column=j, value=v)
    for j in range(1, 5):
        ws3.cell(row=r, column=j).alignment = wrap; ws3.cell(row=r, column=j).border = border
    ws3.row_dimensions[r].height = 58
    r += 1
ws3.merge_cells(f"A{r+1}:D{r+1}")
c = ws3.cell(row=r+1, column=1)
c.value = ("META SEMANAL: +30 ferreterías nuevas · 30 contactadas · 5 clientes nuevos  |  "
           "Redes: +50 seguidores por red · 5 publicaciones · 1 video YouTube")
c.font = Font(bold=True, color="0D0D0D"); c.fill = PatternFill("solid", fgColor=VERDE); c.alignment = wrap
ws3.row_dimensions[r+1].height = 30

# =====================================================================
# HOJA 4 - SEGUIDORES POR RED
# =====================================================================
ws4 = wb.create_sheet("Seguidores por Red")
for col, w in zip("ABC", [16, 60, 48]):
    ws4.column_dimensions[col].width = w
ws4.merge_cells("A1:C1")
c = ws4["A1"]; c.value = "Cómo crecer en cada red — Uniproveedores"
c.font = title_font; c.fill = title_fill; ws4.row_dimensions[1].height = 24
for j, h in enumerate(["Red", "Táctica principal para sumar seguidores", "Frecuencia + tip clave"], start=1):
    cell = ws4.cell(row=2, column=j, value=h)
    cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center; cell.border = border
redes = [
 ("Instagram","Reels cortos (herramienta en acción, precio en pantalla). 3-5 hashtags locales (#ferreteria #herramientas #caba). Respondé DMs con link de WhatsApp. Collabs con proveedores.","1 Reel/día + 3 stories. El gancho en los primeros 2s es lo que trae seguidores nuevos."),
 ("Facebook","Página (no perfil), botón 'Enviar WhatsApp'. Sumate a grupos de compra/venta y de ferreteros. Invitá contactos a seguir la Página.","4-5 posts/semana. Los grupos del rubro son el alcance B2B gratis más rápido."),
 ("Canal WhatsApp","Creá el Canal desde WhatsApp Business. Link en bio IG/FB, firma de mail y en cada venta. Publicá 'ofertas de la semana' y stock nuevo.","2-3 difusiones/semana. Cada cliente que atendés → invitalo. Crece por invitación directa."),
 ("YouTube","Shorts (20-40s) reutilizando Reels + 1 video largo/semana (reviews, 'cómo usar'). Título con keyword + miniatura verde. Pedí suscripción al final.","2-3 Shorts + 1 largo/semana. Shorts = alcance; largos = autoridad y ventas."),
 ("X (Twitter)","Ofertas cortas con imagen, hilos de tips, responder a cuentas del rubro. #Argentina #herramientas.","1-2 posts/día. En X crecés interactuando más que publicando solo."),
 ("LinkedIn","Página de empresa en tono B2B ('mayorista para ferreterías, envío y garantía'). Conectá con dueños/encargados. Publicá casos y catálogo.","2-3 posts/semana. Es la red para VENDERLE a ferreterías, no tanto para volumen."),
]
r = 3
for red, tac, freq in redes:
    ws4.cell(row=r, column=1, value=red).font = Font(bold=True)
    ws4.cell(row=r, column=2, value=tac); ws4.cell(row=r, column=3, value=freq)
    for j in range(1, 4):
        ws4.cell(row=r, column=j).alignment = wrap; ws4.cell(row=r, column=j).border = border
    ws4.row_dimensions[r].height = 72
    r += 1
ws4.merge_cells(f"A{r+1}:C{r+1}")
c = ws4.cell(row=r+1, column=1)
c.value = ("REGLAS PARA TODAS: (1) Mismo @uniproveedoresok y logo. (2) Grabás 1 video → Reel/Short/TikTok/X. "
           "(3) Siempre link a WhatsApp wa.me/541135510715. (4) Horario pico 12-14h y 19-21h. "
           "(5) Respondé todo comentario/DM el mismo día.")
c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=NEGRO); c.alignment = wrap
ws4.row_dimensions[r+1].height = 52

# =====================================================================
# HOJA 5 - MENSAJES DE CONTACTO (plantillas para el Anillo 1)
# =====================================================================
ws5 = wb.create_sheet("Mensajes de contacto")
ws5.column_dimensions["A"].width = 3
ws5.column_dimensions["B"].width = 112

def block5(ws, rows):
    r = 1
    for kind, text in rows:
        cell = ws.cell(row=r, column=2, value=text)
        if kind == "h1":
            cell.font = title_font; cell.fill = title_fill; ws.row_dimensions[r].height = 24
        elif kind == "h2":
            cell.font = Font(bold=True, color="0D0D0D", size=12); cell.fill = PatternFill("solid", fgColor=VERDE)
        elif kind == "msg":
            cell.font = Font(size=10); cell.alignment = wrap
            cell.fill = PatternFill("solid", fgColor="F3F7E6")
            ws.row_dimensions[r].height = 92
        elif kind == "b":
            cell.font = Font(bold=True, size=10); cell.alignment = wrap
        else:
            cell.font = Font(size=10); cell.alignment = wrap
        r += 1

block5(ws5, [
 ("h1", "  Plantillas de contacto — Anillo 1 (Comuna 10)"),
 ("n", "Personalizá lo que está entre [corchetes]. Regla: mensaje corto, valor claro, una sola pregunta y CTA. No mandar catálogo entero en el primer mensaje."),
 ("n", ""),
 ("h2", "  1) WhatsApp — primer mensaje (frío)"),
 ("msg", "Hola, ¿cómo va? Te escribo de UNIPROVEEDORES 🛠️, distribuidora mayorista de herramientas y artículos de ferretería. "
         "Estamos trabajando con ferreterías de [barrio/zona] con precio mayorista, envío y garantía. "
         "¿Con quién puedo hablar por las compras? Te acerco la lista y las ofertas de la semana, sin compromiso. ¡Gracias!"),
 ("h2", "  2) WhatsApp — seguimiento (48-72 hs después, si no respondió)"),
 ("msg", "Hola [nombre]! Te escribí hace unos días de Uniproveedores (mayorista de ferretería). ¿Pudiste ver? "
         "Si querés te paso la lista de precios actualizada de esta semana. Si hay algún producto puntual que más vendés, "
         "decime cuál y te coto al toque 👍"),
 ("h2", "  3) WhatsApp — con oferta gancho (si sabés qué vende)"),
 ("msg", "Hola [nombre]! Esta semana tenemos [producto] a $[precio] mayorista (con envío y garantía). "
         "Es de lo que más rota en ferretería. ¿Te preparo una cotización con cantidades para [nombre de la ferretería]?"),
 ("h2", "  4) Email — presentación"),
 ("b", "  Asunto:  Uniproveedores — precios mayoristas para tu ferretería (envío y garantía)"),
 ("msg", "Hola, ¿cómo estás?\n\nSoy [tu nombre], de UNIPROVEEDORES, distribuidora mayorista de herramientas y artículos de ferretería. "
         "Trabajamos con ferreterías de Capital con precio mayorista, envío y garantía por Mercado Libre.\n\n"
         "Me gustaría acercarte nuestra lista de precios y las ofertas de la semana, sin compromiso. "
         "¿A quién le puedo enviar la información de compras?\n\n"
         "Quedo a disposición.\n[tu nombre] — Uniproveedores\nWhatsApp: 011-3551-0715  ·  ventas@distribuidorauniverso.com\n"
         "uniproveedores.com.ar  ·  Mercado Libre: /tienda/uniproveedores"),
 ("h2", "  5) Guion visita / llamada (para conseguir el nombre del encargado)"),
 ("msg", "Buenas, ¿el encargado de compras se encuentra? Soy de Uniproveedores, distribuidora mayorista de ferretería. "
         "Le quería dejar la lista de precios y ver qué productos le interesan para cotizarle con envío. "
         "¿Con quién tengo el gusto? (→ anotar el nombre en la columna 'Encargado')."),
 ("n", ""),
 ("b", "  Tips de cierre:"),
 ("n", "  • Guardá cada número que contactás como contacto → así ven tu foto/estado de empresa y da confianza."),
 ("n", "  • Después del primer 'sí', mandá SIEMPRE una cotización concreta con cantidades, no una lista genérica."),
 ("n", "  • Invitá a cada contacto a tu Canal de WhatsApp (ofertas) aunque todavía no te compre."),
])

# =====================================================================
# HOJA 6 - TABLERO / RESULTADOS (se actualiza solo con la columna Estado)
# =====================================================================
ws6 = wb.create_sheet("Tablero y Resultados")
ws6.sheet_view.tabColor = VERDE
for col, w in zip("ABCD", [4, 26, 12, 60]):
    ws6.column_dimensions[col].width = w
ws6.merge_cells("A1:D1")
c = ws6["A1"]; c.value = "UNIPROVEEDORES · Tablero de resultados (se actualiza al cargar la columna 'Estado')"
c.font = title_font; c.fill = title_fill; ws6.row_dimensions[1].height = 24

est_range = f"'Ferreterías CABA'!$M$4:$M${len(D)+3}"
pri_range = f"'Ferreterías CABA'!$N$4:$N${len(D)+3}"
ani_range = f"'Ferreterías CABA'!$B$4:$B${len(D)+3}"

ws6.cell(row=3, column=2, value="Embudo de prospección").font = Font(bold=True, size=12)
funnel = ["Sin contactar", "Contactado", "Interesado", "Cotización enviada", "Cliente", "Descartado"]
r = 4
for est in funnel:
    ws6.cell(row=r, column=2, value=est).font = Font(size=10)
    cell = ws6.cell(row=r, column=3, value=f'=COUNTIF({est_range},"{est}")')
    cell.font = Font(bold=True); cell.alignment = center
    cell.fill = PatternFill("solid", fgColor="F3F7E6")
    r += 1
ws6.cell(row=r, column=2, value="TOTAL en la lista").font = Font(bold=True)
ws6.cell(row=r, column=3, value=f'=COUNTA({est_range})').font = Font(bold=True)
ws6.cell(row=r, column=3).alignment = center

# Tasa de conversión
r += 2
ws6.cell(row=r, column=2, value="Tasa de cierre (Cliente / Contactadas+)").font = Font(bold=True, size=11)
# contactadas+ = todo lo que no está 'Sin contactar'
formula_cierre = (f'=IFERROR(COUNTIF({est_range},"Cliente")/'
                  f'(COUNTA({est_range})-COUNTIF({est_range},"Sin contactar")-COUNTIF({est_range},"")),0)')
cc = ws6.cell(row=r, column=3, value=formula_cierre)
cc.number_format = "0%"; cc.font = Font(bold=True); cc.alignment = center

# Por prioridad
r += 2
ws6.cell(row=r, column=2, value="Por prioridad").font = Font(bold=True, size=12); r += 1
for pri in ["Alta", "Media", "Baja"]:
    ws6.cell(row=r, column=2, value=pri).font = Font(size=10)
    ws6.cell(row=r, column=3, value=f'=COUNTIF({pri_range},"{pri}")').alignment = center
    r += 1

# Por anillo
r += 1
ws6.cell(row=r, column=2, value="Por anillo (cercanía al depósito)").font = Font(bold=True, size=12); r += 1
for a in ["Anillo 1", "Anillo 2", "Anillo 3"]:
    ws6.cell(row=r, column=2, value=a).font = Font(size=10)
    ws6.cell(row=r, column=3, value=f'=COUNTIF({ani_range},"{a}")').alignment = center
    r += 1

# Bitácora manual
r += 2
ws6.cell(row=r, column=2, value="Bitácora semanal (anotá a mano)").font = Font(bold=True, size=12); r += 1
for j, h in enumerate(["Semana", "Contactadas", "Notas / aprendizajes"], start=2):
    cell = ws6.cell(row=r, column=j, value=h)
    cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center; cell.border = border
r += 1
for wk in range(1, 9):
    ws6.cell(row=r, column=2, value=f"Semana {wk}").border = border
    ws6.cell(row=r, column=3).border = border
    ws6.cell(row=r, column=4).border = border
    r += 1

out = "/home/user/socialflow/data/Ferreterias-CABA-Uniproveedores.xlsx"
wb.save(out)
print("OK ->", out, "| filas:", len(D))
